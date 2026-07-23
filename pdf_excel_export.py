# pdf_excel_export.py - PipeSight AI
# Supervisor's item #5: "see the output in pdf and the excel sheet
# (structured extracted text)" — renders every matched tag back onto its
# source page image (visual overlay, one combined PDF) and writes a
# structured Excel workbook (Tags / Review Queue / Summary sheets), so a
# reviewer can look at the picture and the spreadsheet side by side and
# visually confirm placement before anything is signed off.
#
# READS THE CANONICAL PER-PAGE JSON (page_json.py), not tags_v2.csv /
# review_queue.csv. Per the "single source of truth" requirement: CSV/PDF/
# Excel/dashboard should all be read-only views generated FROM the JSON,
# not independent queries that can drift out of sync with human review
# decisions. The JSON is written by tag_extractor_v2.py after each run and
# re-written by review_server.py after every human review action, so this
# script always sees current state, including corrections/promotions a
# human made after the pipeline ran.
#
# Exception: dropped_candidates.csv stays a separate, CSV-only audit trail.
# It's never written to the DB or reviewable — nothing to keep in sync,
# since a dropped candidate is never edited by anyone. Loaded directly if
# present; harmless if absent.
#
# Usage:
#   python pdf_excel_export.py --json-dir data\structured_output
#   python pdf_excel_export.py --json-dir data\structured_output --page "SET 1_page_1"
#   python pdf_excel_export.py --json-dir data\structured_output --no-pdf   (Excel only)
#   python pdf_excel_export.py --json-dir data\structured_output --no-excel (PDF only)
#
# Output:
#   data\tag_extraction_v2\pipesight_review.pdf   (overlay, one page per source page)
#   data\tag_extraction_v2\pipesight_review.xlsx  (Tags / Review Queue / Dropped / Summary)

import csv
import json
import argparse
from pathlib import Path
from collections import defaultdict

from PIL import Image, ImageDraw, ImageFont

OUT_DIR = Path("data") / "tag_extraction_v2"
JSON_DIR = Path("data") / "structured_output"

DROPPED_CSV   = OUT_DIR / "dropped_candidates.csv"   # audit-only, see note above

PDF_OUT  = OUT_DIR / "pipesight_review.pdf"
XLSX_OUT = OUT_DIR / "pipesight_review.xlsx"

# Longest side any overlay page image gets downscaled to before going into
# the PDF. Source pages run ~10000px+ on the long side (see handoff:
# 9934x7017) — embedding at full resolution makes a multi-page PDF
# enormous and slow to open for no reviewer-visible benefit at normal zoom.
MAX_OVERLAY_DIM = 2200

# ─── STATUS -> COLOR ──────────────────────────────────────────────────────
# One color per what a reviewer actually needs to triage at a glance:
# clean auto-accepted tags vs. anything that still wants a human look.
COLOR_VALID_AUTO      = (34, 139, 34)     # forest green  — auto_accept=True
COLOR_VALID_REVIEW    = (218, 165, 32)    # goldenrod     — VALID but low-trust/HUMAN_REQUIRED
COLOR_RESCUED         = (148, 0, 211)     # purple        — recovered via YOLO rescue (class_id=-1)
COLOR_RECOVERED_ROT   = (30, 144, 255)    # dodger blue   — RECOVERED_ROTATED
COLOR_RAW             = (255, 140, 0)     # orange        — RAW (not ISA-shaped)
COLOR_MISSING         = (220, 20, 60)     # crimson       — MISSING (no text found at all)
COLOR_REVIEW_QUEUE    = (105, 105, 105)   # dim gray      — orphaned/misrouted candidate, dashed

LEGEND = [
    ("Auto-accepted (VALID)",            COLOR_VALID_AUTO),
    ("VALID — needs human sign-off",     COLOR_VALID_REVIEW),
    ("Rescued (YOLO low-conf recovery)", COLOR_RESCUED),
    ("Recovered (rotated text)",         COLOR_RECOVERED_ROT),
    ("RAW (not ISA-shaped)",             COLOR_RAW),
    ("MISSING (no text found)",          COLOR_MISSING),
    ("Review queue candidate",           COLOR_REVIEW_QUEUE),
]

# ── Non-instrument YOLO detections (--show-all-detections) ──
# Thin, light, class-labeled boxes so the full model output is visible on
# request without competing with the bold instrument/tag boxes above.
COLOR_DET_MECHANICAL = (100, 149, 237)   # cornflower blue — valves/pumps/vessels/etc
COLOR_DET_STRUCTURAL  = (160, 120, 90)   # tan/brown       — reducers, flanges, flow arrows
COLOR_DET_REFERENCE   = (150, 150, 150)  # mid gray        — "Box" label/reference callouts
COLOR_DET_UNKNOWN     = (40, 40, 40)     # near-black      — unrecognized class_id

ROUTING_DET_COLOR = {
    "mechanical": COLOR_DET_MECHANICAL,
    "structural": COLOR_DET_STRUCTURAL,
    "reference":  COLOR_DET_REFERENCE,
    "unknown_class": COLOR_DET_UNKNOWN,
}

LEGEND_ALL_DETECTIONS = [
    ("Detection: mechanical (valve/pump/vessel/etc)", COLOR_DET_MECHANICAL),
    ("Detection: structural (reducer/flange/etc)",     COLOR_DET_STRUCTURAL),
    ("Detection: reference/label box",                 COLOR_DET_REFERENCE),
    ("Detection: unrecognized class",                  COLOR_DET_UNKNOWN),
]

# ─── CSV LOADING ─────────────────────────────────────────────────────────

def load_csv_rows(path):
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))

def load_page_json_files(json_dir, page=None):
    """
    Reads every {page}_tags.json under json_dir (recursively, since pages
    live one-per-SET-folder) and flattens page_json.py's {tag: {...}}
    schema back into the flat per-row dicts this file's drawing/Excel code
    already expects — same field names as the old tags_v2.csv/
    review_queue.csv had, so row_color_and_label/build_excel don't need to
    change. Returns (tags_rows, review_rows, page_names) in first-seen
    (i.e. folder/file sort) order.
    """
    json_dir = Path(json_dir)
    paths = sorted(json_dir.rglob("*_tags.json"))
    if page:
        paths = [p for p in paths if p.stem == f"{page}_tags"]

    tags_rows, review_rows, page_names = [], [], []

    for p in paths:
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as e:
            print(f"  WARNING: skipping unreadable {p}: {e}")
            continue

        page_name = data.get("page")
        if not page_name:
            continue
        page_names.append(page_name)

        for tag_text, entry in data.get("tags", {}).items():
            td = entry.get("technical_details", {})
            loc = entry.get("location", {})
            # UNREAD_<det_id> (and its rare __det<id> dedup suffix) is a
            # synthetic key for MISSING tags, never real tag text.
            display_tag = "" if tag_text.startswith("UNREAD_") else tag_text
            tags_rows.append({
                "page": page_name,
                "det_id": td.get("det_id"),
                "class_name": td.get("class_name"),
                "class_id": td.get("class_id"),
                "yolo_conf": td.get("detection_confidence"),
                "x1": loc.get("x1"), "y1": loc.get("y1"),
                "x2": loc.get("x2"), "y2": loc.get("y2"),
                "routing": "instrument",
                "raw_ocr": td.get("raw_reading") or "",
                "ocr_conf": td.get("reading_confidence"),
                "tag": display_tag,
                "tag_status": td.get("tag_status", ""),
                "inferred_function": entry.get("instrument_type") or "",
                "combined_conf": td.get("combined_confidence"),
                "auto_accept": entry.get("review_status") == "auto_confirmed",
                "verified_by": entry.get("reviewed_by") or "",
                "verified_at": entry.get("reviewed_at") or "",
                "verification_method": td.get("verification_method") or "",
            })

        for sym in data.get("non_instrument_symbols", []):
            td = sym.get("technical_details", {})
            loc = sym.get("location", {})
            tags_rows.append({
                "page": page_name,
                "det_id": td.get("det_id"),
                "class_name": sym.get("class_name"),
                "class_id": None,
                "yolo_conf": td.get("detection_confidence"),
                "x1": loc.get("x1"), "y1": loc.get("y1"),
                "x2": loc.get("x2"), "y2": loc.get("y2"),
                "routing": sym.get("routing"),
                "raw_ocr": "", "ocr_conf": 0.0, "tag": "",
                "tag_status": "SKIPPED", "inferred_function": "",
                "combined_conf": 0.0, "auto_accept": False,
                "verified_by": "", "verified_at": "", "verification_method": "",
            })

        for cand in data.get("review_queue", []):
            td = cand.get("technical_details", {})
            if td.get("status") != "OPEN":
                continue  # resolved/dismissed — don't render as pending review
            loc = cand.get("location", {})
            review_rows.append({
                "page": page_name,
                "kind": cand.get("kind", ""),
                "candidate_tag": cand.get("candidate_tag") or "",
                "inferred_function": cand.get("instrument_type") or "",
                "raw_text": cand.get("raw_reading") or "",
                "ocr_conf": td.get("reading_confidence"),
                "x1": loc.get("x1"), "y1": loc.get("y1"),
                "x2": loc.get("x2"), "y2": loc.get("y2"),
                "nearest_class_name": td.get("nearest_class_name"),
                "nearest_routing": td.get("nearest_routing"),
                "nearest_dist_px": td.get("nearest_distance_px"),
            })

    return tags_rows, review_rows, page_names

def to_float(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default

def to_int(v, default=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default

def to_bool(v):
    return str(v).strip().lower() in ("true", "1", "yes")

def row_color_and_label(row, show_all_detections=False):
    """
    Decide overlay box color + short label for one tags_v2.csv row.
    Precedence: rescued > recovered-rotated > raw/missing > valid variants.

    Non-instrument (SKIPPED) rows are only drawn when show_all_detections
    is True — every YOLO detection on the page (mechanical/structural/
    reference/unknown), thin and class-labeled, distinct from the bold
    instrument/tag boxes so the two don't get visually confused.
    """
    tag_status = row.get("tag_status", "")
    tag = row.get("tag") or row.get("candidate_tag") or ""
    is_rescued = to_int(row.get("class_id"), 0) == -1

    if tag_status == "SKIPPED":
        if not show_all_detections:
            return None, None
        color = ROUTING_DET_COLOR.get(row.get("routing"), COLOR_DET_UNKNOWN)
        return color, row.get("class_name", "")

    if is_rescued:
        return COLOR_RESCUED, f"{tag} (RESCUED)" if tag else "RESCUED"
    if tag_status == "RECOVERED_ROTATED":
        return COLOR_RECOVERED_ROT, f"{tag} (rotated)" if tag else "RECOVERED"
    if tag_status == "RAW":
        return COLOR_RAW, tag or "(unreadable)"
    if tag_status == "MISSING":
        return COLOR_MISSING, "MISSING"
    if tag_status == "VALID":
        if to_bool(row.get("auto_accept")):
            return COLOR_VALID_AUTO, tag
        if row.get("verified_by"):
            return COLOR_VALID_AUTO, f"{tag} (confirmed)"   # or a distinct color if you want auto vs human visually different
        return COLOR_VALID_REVIEW, f"{tag} (verify)"
    return (128, 128, 128), tag or tag_status

# ─── FONT ─────────────────────────────────────────────────────────────────

def load_font(size):
    for candidate in (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ):
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()

# ─── OVERLAY DRAWING ──────────────────────────────────────────────────────

def draw_dashed_rect(draw, box, color, width=3, dash=14, gap=8):
    x1, y1, x2, y2 = box
    for (sx, sy, ex, ey) in [(x1, y1, x2, y1), (x2, y1, x2, y2),
                             (x2, y2, x1, y2), (x1, y2, x1, y1)]:
        length = max(abs(ex - sx), abs(ey - sy))
        steps = max(1, int(length / (dash + gap)))
        for i in range(steps + 1):
            t0 = i * (dash + gap) / max(length, 1)
            t1 = min(1.0, t0 + dash / max(length, 1))
            px0, py0 = sx + (ex - sx) * t0, sy + (ey - sy) * t0
            px1, py1 = sx + (ex - sx) * t1, sy + (ey - sy) * t1
            draw.line([(px0, py0), (px1, py1)], fill=color, width=width)

def draw_label(draw, x, y, text, color, font, page_w, page_h):
    if not text:
        return
    bbox = draw.textbbox((0, 0), text, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    pad = 3
    lx0, ly0 = x, max(0, y - th - 2 * pad)
    lx1, ly1 = min(page_w, x + tw + 2 * pad), ly0 + th + 2 * pad
    draw.rectangle([lx0, ly0, lx1, ly1], fill=color)
    draw.text((lx0 + pad, ly0 + pad), text, fill=(255, 255, 255), font=font)

def draw_legend(draw, page_w, page_h, font, show_all_detections=False):
    entries = LEGEND + (LEGEND_ALL_DETECTIONS if show_all_detections else [])
    x0, y0 = 20, page_h - (len(entries) * 26) - 20
    draw.rectangle([x0 - 10, y0 - 10, x0 + 400, page_h - 10],
                    fill=(255, 255, 255), outline=(0, 0, 0), width=2)
    for i, (label, color) in enumerate(entries):
        y = y0 + i * 26
        draw.rectangle([x0, y, x0 + 20, y + 16], fill=color)
        draw.text((x0 + 28, y - 2), label, fill=(0, 0, 0), font=font)

def build_overlay_image(image_path, tag_rows, review_rows, page_label,
                         font_label, font_legend, show_all_detections=False):
    img = Image.open(image_path).convert("RGB")
    orig_w, orig_h = img.size
    scale = min(1.0, MAX_OVERLAY_DIM / max(orig_w, orig_h))
    if scale < 1.0:
        img = img.resize((max(1, int(orig_w * scale)), max(1, int(orig_h * scale))), Image.LANCZOS)
    page_w, page_h = img.size
    draw = ImageDraw.Draw(img)

    # Draw ALL YOLO detections first (thin, unlabeled — this is a sanity-
    # check layer, not meant to be read tag-by-tag) so the bold instrument/
    # tag boxes always render on top and stay the visually dominant layer.
    if show_all_detections:
        for row in tag_rows:
            if row.get("tag_status") != "SKIPPED":
                continue
            color = ROUTING_DET_COLOR.get(row.get("routing"), COLOR_DET_UNKNOWN)
            x1 = to_int(row["x1"]) * scale
            y1 = to_int(row["y1"]) * scale
            x2 = to_int(row["x2"]) * scale
            y2 = to_int(row["y2"]) * scale
            draw.rectangle([x1, y1, x2, y2], outline=color, width=1)
            class_name = row.get("class_name", "")
            if class_name:
                draw_label(draw, x1, y1, class_name, color, font_legend, page_w, page_h)

    for row in tag_rows:
        color, label = row_color_and_label(row)
        if color is None:
            continue
        x1 = to_int(row["x1"]) * scale
        y1 = to_int(row["y1"]) * scale
        x2 = to_int(row["x2"]) * scale
        y2 = to_int(row["y2"]) * scale
        draw.rectangle([x1, y1, x2, y2], outline=color, width=4)
        draw_label(draw, x1, y1, label, color, font_label, page_w, page_h)

    for row in review_rows:
        x1 = to_int(row["x1"]) * scale
        y1 = to_int(row["y1"]) * scale
        x2 = to_int(row["x2"]) * scale
        y2 = to_int(row["y2"]) * scale
        draw_dashed_rect(draw, (x1, y1, x2, y2), COLOR_REVIEW_QUEUE, width=3)
        label = row.get("candidate_tag") or row.get("kind") or "REVIEW"
        draw_label(draw, x1, y1, f"{label}?", COLOR_REVIEW_QUEUE, font_label, page_w, page_h)

    # Page title bar
    draw.rectangle([0, 0, page_w, 34], fill=(20, 20, 20))
    title = page_label + ("  [all detections shown]" if show_all_detections else "")
    draw.text((10, 6), title, fill=(255, 255, 255), font=font_label)

    draw_legend(draw, page_w, page_h, font_legend, show_all_detections=show_all_detections)
    return img

def resolve_image_path(page_name, images_dir):
    direct = images_dir / f"{page_name}.png"
    if direct.exists():
        return direct
    matches = list(images_dir.rglob(f"{page_name}.png"))
    return matches[0] if matches else None

def build_overlay_pdf(page_names, tags_by_page, review_by_page, images_dir, out_path,
                       show_all_detections=False):
    """
    Assembles one PDF page per source page using reportlab, embedding each
    overlay image full-bleed at its own pixel aspect ratio (1px -> 1pt, so
    a 2200px-wide overlay becomes a 2200pt-wide PDF page — reportlab has
    no fixed page-size requirement, unlike letter/A4-only tools).

    Deliberately NOT using PIL's Image.save(..., "PDF") here: that path
    requires Pillow's own JPEG encoder plugin to be registered, and on at
    least one real Windows venv (confirmed) that plugin was missing/broken
    (KeyError: 'JPEG') even though PNG save/open worked fine. reportlab's
    drawImage() embeds the same in-memory PNG data directly and never
    touches that codepath, so this failure mode can't recur here.
    """
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.utils import ImageReader
    except ModuleNotFoundError:
        print("  ERROR: the 'reportlab' package is not installed in this environment.\n"
              "  Install it with:  pip install reportlab\n"
              "  Then re-run this command. (Skipping overlay PDF for now — "
              "Excel export below is unaffected.)")
        return None

    font_label  = load_font(20)
    font_legend = load_font(16)
    missing_images = []
    built_any = False
    c = None

    for i, page_name in enumerate(page_names, 1):
        img_path = resolve_image_path(page_name, images_dir)
        if img_path is None:
            missing_images.append(page_name)
            continue
        label = f"Page {i}/{len(page_names)}: {page_name}"
        overlay = build_overlay_image(
            img_path, tags_by_page.get(page_name, []),
            review_by_page.get(page_name, []), label, font_label, font_legend,
            show_all_detections=show_all_detections,
        )
        w, h = overlay.size
        if c is None:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            c = rl_canvas.Canvas(str(out_path), pagesize=(w, h))
        else:
            c.setPageSize((w, h))
        c.drawImage(ImageReader(overlay), 0, 0, width=w, height=h)
        c.showPage()
        built_any = True

    if not built_any:
        print("  No page images found — overlay PDF NOT written. "
              "Check --images-dir points at the same folder tag_extractor_v2.py used.")
        return None

    c.save()
    print(f"  Overlay PDF written: {out_path}  ({len(page_names) - len(missing_images)} page(s))")
    if missing_images:
        print(f"  WARNING: image not found for {len(missing_images)} page(s) "
              f"(skipped in PDF): {', '.join(missing_images)}")
    return out_path

# ─── EXCEL WORKBOOK ───────────────────────────────────────────────────────

def build_excel(tags_rows, review_rows, dropped_rows, page_names, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial")
    BODY_FONT   = Font(name="Arial")
    STATUS_FILL = {
        "VALID_AUTO":   PatternFill("solid", fgColor="C6EFCE"),
        "VALID_REVIEW": PatternFill("solid", fgColor="FFEB9C"),
        "RESCUED":      PatternFill("solid", fgColor="E4C6F1"),
        "RECOVERED_ROTATED": PatternFill("solid", fgColor="BDD7EE"),
        "RAW":          PatternFill("solid", fgColor="FCE4D6"),
        "MISSING":      PatternFill("solid", fgColor="F8CBAD"),
    }

    def style_header(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def autosize(ws, headers, rows_as_lists):
        for c, h in enumerate(headers, 1):
            width = max(len(str(h)), *(len(str(r[c-1])) for r in rows_as_lists)) if rows_as_lists else len(str(h))
            ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 45)

    # Page number lookup for cross-referencing with the overlay PDF
    page_number = {p: i + 1 for i, p in enumerate(page_names)}

    # ── Sheet: Tags (instrument symbols only — the actual deliverable) ──
    ws = wb.active
    ws.title = "Tags"
    headers = ["overlay_pdf_page", "page", "det_id", "class_name", "tag", "tag_status",
               "verification_method", "auto_accept", "combined_conf", "yolo_conf",
               "ocr_conf", "raw_ocr", "inferred_function",
               "x1", "y1", "x2", "y2"]
    style_header(ws, headers)
    body_rows = []
    r = 2
    for row in tags_rows:
        if row.get("routing") != "instrument":
            continue
        is_rescued = to_int(row.get("class_id"), 0) == -1
        auto_accept = to_bool(row.get("auto_accept"))
        status = row.get("tag_status", "")
        values = [
            page_number.get(row["page"], ""), row["page"], row["det_id"], row["class_name"],
            row.get("tag", ""), status, row.get("verification_method", ""),
            "Yes" if auto_accept else "No", to_float(row.get("combined_conf")),
            to_float(row.get("yolo_conf")), to_float(row.get("ocr_conf")),
            row.get("raw_ocr", ""), row.get("inferred_function", ""),
            row.get("x1", ""), row.get("y1", ""), row.get("x2", ""), row.get("y2", ""),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
        # Row fill by status, matching overlay PDF colors so the two line up visually
        if is_rescued:
            fill = STATUS_FILL["RESCUED"]
        elif status == "RECOVERED_ROTATED":
            fill = STATUS_FILL["RECOVERED_ROTATED"]
        elif status == "RAW":
            fill = STATUS_FILL["RAW"]
        elif status == "MISSING":
            fill = STATUS_FILL["MISSING"]
        elif status == "VALID":
            fill = STATUS_FILL["VALID_AUTO"] if auto_accept else STATUS_FILL["VALID_REVIEW"]
        else:
            fill = None
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = fill
        body_rows.append(values)
        r += 1
    autosize(ws, headers, body_rows)

    # ── Sheet: Review Queue ──
    ws2 = wb.create_sheet("Review Queue")
    rq_headers = ["overlay_pdf_page", "page", "kind", "candidate_tag", "inferred_function",
                  "raw_text", "ocr_conf", "nearest_class_name", "nearest_routing",
                  "nearest_dist_px", "x1", "y1", "x2", "y2"]
    style_header(ws2, rq_headers)
    rq_body = []
    r = 2
    for row in review_rows:
        values = [page_number.get(row["page"], ""), row["page"], row.get("kind", ""),
                  row.get("candidate_tag", ""), row.get("inferred_function", ""),
                  row.get("raw_text", ""), to_float(row.get("ocr_conf")),
                  row.get("nearest_class_name", ""), row.get("nearest_routing", ""),
                  to_float(row.get("nearest_dist_px")),
                  row.get("x1", ""), row.get("y1", ""), row.get("x2", ""), row.get("y2", "")]
        for c, v in enumerate(values, 1):
            ws2.cell(row=r, column=c, value=v).font = BODY_FONT
        rq_body.append(values)
        r += 1
    autosize(ws2, rq_headers, rq_body)

    # ── Sheet: Dropped Candidates (only if present — audit trail) ──
    if dropped_rows:
        ws3 = wb.create_sheet("Dropped Candidates")
        dc_headers = rq_headers + ["drop_reason"]
        style_header(ws3, dc_headers)
        dc_body = []
        r = 2
        for row in dropped_rows:
            values = [page_number.get(row["page"], ""), row["page"], row.get("kind", ""),
                      row.get("candidate_tag", ""), row.get("inferred_function", ""),
                      row.get("raw_text", ""), to_float(row.get("ocr_conf")),
                      row.get("nearest_class_name", ""), row.get("nearest_routing", ""),
                      to_float(row.get("nearest_dist_px")),
                      row.get("x1", ""), row.get("y1", ""), row.get("x2", ""), row.get("y2", ""),
                      row.get("drop_reason", "")]
            for c, v in enumerate(values, 1):
                ws3.cell(row=r, column=c, value=v).font = BODY_FONT
            dc_body.append(values)
            r += 1
        autosize(ws3, dc_headers, dc_body)

    # ── Sheet: Summary (per page) ──
    ws4 = wb.create_sheet("Summary")
    sm_headers = ["overlay_pdf_page", "page", "instrument_symbols", "valid", "valid_pct",
                  "raw", "missing", "recovered_rotated", "rescued", "review_queue_items"]
    style_header(ws4, sm_headers)

    per_page = defaultdict(lambda: {"inst": 0, "valid": 0, "raw": 0, "missing": 0,
                                     "recovered": 0, "rescued": 0})
    for row in tags_rows:
        if row.get("routing") != "instrument":
            continue
        p = row["page"]
        per_page[p]["inst"] += 1
        status = row.get("tag_status", "")
        if status == "VALID":
            per_page[p]["valid"] += 1
        elif status == "RAW":
            per_page[p]["raw"] += 1
        elif status == "MISSING":
            per_page[p]["missing"] += 1
        elif status == "RECOVERED_ROTATED":
            per_page[p]["recovered"] += 1
        if to_int(row.get("class_id"), 0) == -1:
            per_page[p]["rescued"] += 1
    review_counts = Counter = defaultdict(int)
    for row in review_rows:
        review_counts[row["page"]] += 1

    sm_body = []
    r = 2
    for p in page_names:
        s = per_page.get(p, {"inst": 0, "valid": 0, "raw": 0, "missing": 0,
                              "recovered": 0, "rescued": 0})
        valid_pct = round(s["valid"] / s["inst"] * 100, 1) if s["inst"] else 0.0
        values = [page_number.get(p, ""), p, s["inst"], s["valid"], valid_pct,
                  s["raw"], s["missing"], s["recovered"], s["rescued"], review_counts.get(p, 0)]
        for c, v in enumerate(values, 1):
            ws4.cell(row=r, column=c, value=v).font = BODY_FONT
        sm_body.append(values)
        r += 1
    autosize(ws4, sm_headers, sm_body)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Excel workbook written: {out_path}  "
          f"(sheets: {', '.join(ws.title for ws in wb.worksheets)})")
    return out_path

# ─── MAIN ─────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="PipeSight AI — render tag_extractor_v2.py's results into a "
                     "visual PDF overlay + structured Excel workbook for human review.")
    ap.add_argument("--json-dir", type=str, default=str(JSON_DIR),
                     help="Root of the canonical per-page JSON output "
                          "(same folder page_json.py writes to), e.g. data\\structured_output")
    ap.add_argument("--dropped-csv", type=str, default=str(DROPPED_CSV),
                     help="Audit-only CSV of filtered-out candidates (not part of "
                          "the canonical JSON — see module docstring). Optional.")
    ap.add_argument("--images-dir", type=str, required=True,
                     help="Root images folder (same one passed to tag_extractor_v2.py), "
                          "e.g. data\\haifa_real_pids\\images")
    ap.add_argument("--page", type=str, default=None,
                     help="Only include this page (default: every page JSON found)")
    ap.add_argument("--pdf-out", type=str, default=str(PDF_OUT))
    ap.add_argument("--xlsx-out", type=str, default=str(XLSX_OUT))
    ap.add_argument("--no-pdf", action="store_true", help="Skip building the overlay PDF")
    ap.add_argument("--no-excel", action="store_true", help="Skip building the Excel workbook")
    ap.add_argument("--show-all-detections", action="store_true",
                     help="(Kept for backward compatibility — this is now the default; "
                          "use --no-all-detections to turn it off.)")
    ap.add_argument("--no-all-detections", action="store_true",
                     help="Skip drawing non-instrument YOLO detections (mechanical/"
                          "structural/reference/unknown) as thin, class-labeled, "
                          "colored boxes underneath the bold instrument/tag boxes. "
                          "By default these ARE drawn so you can see every identified "
                          "class alongside the tag-extraction results.")
    args = ap.parse_args()

    json_dir_path = Path(args.json_dir)
    if not json_dir_path.exists():
        print(f"JSON output folder not found at {json_dir_path}. "
              f"Run tag_extractor_v2.py first (it writes page JSON as part of its DB write).")
        return

    tags_rows, review_rows, page_names = load_page_json_files(json_dir_path, page=args.page)
    dropped_rows = load_csv_rows(Path(args.dropped_csv))
    if args.page:
        dropped_rows = [r for r in dropped_rows if r["page"] == args.page]

    if not tags_rows:
        print(f"No tag data found{' for page ' + args.page if args.page else ''} "
              f"under {json_dir_path}.")
        return

    tags_by_page = defaultdict(list)
    for row in tags_rows:
        tags_by_page[row["page"]].append(row)
    review_by_page = defaultdict(list)
    for row in review_rows:
        review_by_page[row["page"]].append(row)

    print(f"\n{'='*60}")
    print(f"  PipeSight AI — PDF + Excel Review Export")
    print(f"{'='*60}")
    print(f"  Source         : {json_dir_path}")
    print(f"  Pages          : {len(page_names)}")
    print(f"  Tag rows       : {len(tags_rows)}")
    print(f"  Review rows    : {len(review_rows)}")
    print(f"  Dropped rows   : {len(dropped_rows)}")
    print(f"{'='*60}\n")

    if not args.no_pdf:
        build_overlay_pdf(page_names, tags_by_page, review_by_page,
                           Path(args.images_dir), Path(args.pdf_out),
                           show_all_detections=not args.no_all_detections)
    if not args.no_excel:
        build_excel(tags_rows, review_rows, dropped_rows, page_names, Path(args.xlsx_out))

    print(f"\n{'='*60}")
    print(f"  Done.")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
